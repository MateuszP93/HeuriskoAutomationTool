from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from heurisko_automation.config import read_yaml


class QueueRunner:
    def __init__(self, workflow_runner):
        self.workflow_runner = workflow_runner

    def run_yaml(self, path: Path):
        data = read_yaml(path)
        results = []
        for item in data.get("queue", []):
            results.append(self._run_item(item))
        return results

    def run_excel(self, path: Path):
        workbook = load_workbook(path, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(cell).strip() for cell in rows[0]]
        results = []
        for row in rows[1:]:
            record = dict(zip(headers, row))
            workflow = record.pop("workflow")
            run_id = record.pop("run_id")
            params = {key: value for key, value in record.items() if value is not None}
            results.append(
                self._run_item({"workflow": workflow, "run_id": run_id, "params": params})
            )
        return results

    def _run_item(self, item: dict[str, Any]):
        params = dict(item.get("params", {}))
        params["run_id"] = item["run_id"]
        return self.workflow_runner.run_workflow(item["workflow"], params)
